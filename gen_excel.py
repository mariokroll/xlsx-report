import pandas as pd
import matplotlib.pyplot as plt
import openpyxl as op
from openpyxl.chart import Reference, DoughnutChart


def extract() -> pd.DataFrame:
    """
    Extract the data from the csv file
    """
    prediction = pd.read_csv('2017_prediction.csv')
    prediction.set_index('Unnamed: 0', inplace=True)
    pizzas2 = pd.read_csv('df_merged.csv')
    pizzas = pd.read_csv('pizzas.csv')
    return prediction, pizzas2, pizzas


def transform(pizzas2: pd.DataFrame, pizzas: pd.DataFrame) -> pd.DataFrame:
    """
    Transform the csv, including new columns that will be useful for the analysis and plots
    """
    df = pd.merge(pizzas2, pizzas.loc[:, ['pizza_id', 'price']], on='pizza_id')
    df['price'] = df['price'] * df['quantity']
    df['date'] = pd.to_datetime(df['date'])
    df["pizza_name"] = df['pizza_id'].str.replace(r'_[a-z]+$', '', regex=True)
    df['month'] = df['date'].dt.month
    sold_pizzas = df['pizza_name'].value_counts()
    sizes = df['pizza_id'].str.replace(r'[a-z]+_', '', regex=True).value_counts()
    return df, sold_pizzas, sizes


def create_images(df: pd.DataFrame, prediction: pd.DataFrame):
    """
    Create the images that we are going to paste in the xlsx file
    """

    names = []

    df.groupby('week')['price'].sum().plot(kind='line', figsize=(10, 5), title='Annual profit per week', color='black')
    plt.xlabel('Week')
    plt.ylabel('Profit ($)')
    plt.grid()
    plt.savefig('Annual_profit.png', bbox_inches='tight')
    plt.close()

    df.groupby('pizza_name')['quantity'].sum().sort_values(ascending=False).plot(kind='bar', figsize=(10, 5), title='Pizzas sold in a year')
    plt.ylabel('Quantity')
    plt.xlabel('Pizza')
    plt.savefig('Pizzas_sold.png', bbox_inches='tight')
    plt.close()

    df.groupby('pizza_name')['price'].sum().sort_values(ascending=False).plot(kind='bar', figsize=(10, 5), title='Most profitable pizzas')
    plt.ylabel('Profit ($)')
    plt.xlabel('Pizza')
    plt.savefig('Most_profitable_pizzas.png', bbox_inches='tight')
    plt.close()

    df.groupby('month')['quantity'].sum().plot(kind='line', figsize=(10, 5), title='Monthly amount of pizzas sold', color='orange')
    plt.grid()
    plt.xlabel('Month')
    plt.ylabel('Quantity')
    plt.savefig('Pizzas_sold_month.png', bbox_inches='tight')
    plt.close()

    for row in prediction.iterrows():
        fig = plt.figure(figsize=(10, 5))
        plt.bar(row[1].index, row[1].values, color='green')
        plt.xticks(rotation=90)
        plt.title('Ingredients used on {}'.format(row[0]))
        plt.xlabel('Ingredients')
        plt.ylabel('Quantity')
        plt.savefig('image5_{}.png'.format(row[0]), bbox_inches='tight')


def create_xlsx_file(sizes):
    wb = op.Workbook()
    ws = wb.active
    ws.title = 'Reporte ejecutivo'
    # Get the images
    img1 = op.drawing.image.Image('Annual_profit.png')
    img2 = op.drawing.image.Image('Most_profitable_pizzas.png')
    img3 = op.drawing.image.Image('Pizzas_sold.png')
    img4 = op.drawing.image.Image('Pizzas_sold_month.png')
    imgs5 = [op.drawing.image.Image('image5_week {}.png'.format(i)) for i in range(1, 53)]

    # Add the images to the xlsx file
    img1.height = 280
    img1.width = 520
    ws.add_image(img1, 'B2')

    img2.height = 280
    img2.width = 520
    ws.add_image(img2, 'K2')

    img3.height = 280
    img3.width = 520
    limit_columns = 1
    contador = 0
    ad = 25

    ws1 = wb.create_sheet('Reporte de ingredientes')
    for im in imgs5:
        if contador % 2 == 0:
            im.height = 400
            im.width = 620
            ws1.add_image(im, 'B{}'.format(limit_columns))
            contador += 1
        else:
            im.height = 400
            im.width = 620
            ws1.add_image(im, 'N{}'.format(limit_columns))
            limit_columns += 27
            contador += 1

    ws2 = wb.create_sheet('Reporte de pedidos')
    img3.height = 280
    img3.width = 520
    img4.height = 280
    img4.width = 520


    # Creamos el gráfico de disco
    for row in sizes.iteritems():
        ws2.append(row)
    chart = DoughnutChart()
    chart.title = 'Pizza Sales'
    chart.height = 10
    chart.width = 10
    chart.style = 10
    labels = Reference(ws2, min_col=1, min_row=1, max_row=5)
    data = Reference(ws2, min_col=2, min_row=1, max_row=5)
    chart.add_data(data)
    chart.set_categories(labels)
    ws2.add_chart(chart, 'J2')
    ws2.add_image(img4, 'A8')
    ws2.add_image(img3, 'A24')

    # Save the file
    wb.save('reporte.xlsx')


if __name__ == '__main__':
    prediction, pizzas2, pizzas = extract()
    df, sold_pizzas, sizes = transform(pizzas2, pizzas)
    create_images(df, prediction)
    create_xlsx_file(sizes)
